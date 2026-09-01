from __future__ import annotations

import re
import struct

from pathlib import Path
from urllib.parse import urlparse
from urllib.request import (
    Request,
    urlopen,
)

from django.core.management.base import (
    BaseCommand,
    CommandError,
)
from django.db import transaction

from openpyxl import load_workbook

from apps.core.models import (
    AntivibrationTechnicalData,
)


def exact_key(value):
    return str(
        value or ""
    ).strip().casefold()


def get_source(cell):

    if (
        cell.hyperlink
        and cell.hyperlink.target
    ):
        return str(
            cell.hyperlink.target
        ).strip()

    value = str(
        cell.value or ""
    ).strip()

    match = re.match(
        r'^\s*=HYPERLINK\(\s*"([^"]+)"',
        value,
        re.IGNORECASE,
    )

    if match:
        return match.group(1).strip()

    return value


def validate_source_url(
    source_url,
):

    parsed = urlparse(
        source_url
    )

    if (
        parsed.scheme
        not in {
            "http",
            "https",
        }
        or not parsed.hostname
    ):
        raise ValueError(
            "La fuente 3D no es "
            "una URL HTTP/HTTPS válida."
        )


def validate_remote_glb(
    source_url,
):

    request = Request(
        source_url,
        headers={
            "User-Agent": (
                "SMAV-INAHER-"
                "3D-Sync/1.0"
            ),
            "Accept": (
                "model/gltf-binary,"
                "application/octet-stream,"
                "*/*"
            ),
            "Range": "bytes=0-11",
        },
    )

    with urlopen(
        request,
        timeout=30,
    ) as response:

        header = response.read(
            12
        )

        content_type = (
            response.headers.get(
                "Content-Type",
                "",
            )
            .split(";")[0]
            .strip()
        )


    if len(header) < 12:
        raise ValueError(
            "El origen no devolvió "
            "cabecera GLB."
        )


    magic, version, _ = (
        struct.unpack(
            "<III",
            header,
        )
    )


    if magic != 0x46546C67:
        raise ValueError(
            "La URL no contiene "
            "un archivo GLB."
        )


    if version != 2:
        raise ValueError(
            f"GLB versión {version}; "
            "se requiere glTF 2.0."
        )


    if content_type not in {
        "model/gltf-binary",
        "application/octet-stream",
        "binary/octet-stream",
    }:
        raise ValueError(
            "Content-Type inesperado: "
            f"{content_type}"
        )


class Command(BaseCommand):

    help = (
        "Guarda en PostgreSQL las URL "
        "de los modelos 3D del Excel."
    )


    def add_arguments(
        self,
        parser,
    ):

        parser.add_argument(
            "--excel",
            required=True,
        )

        parser.add_argument(
            "--sheet",
            default="Antivibratorios",
        )

        parser.add_argument(
            "--dry-run",
            action="store_true",
        )

        parser.add_argument(
            "--apply",
            action="store_true",
        )


    def handle(
        self,
        *args,
        **options,
    ):

        if (
            options["dry_run"]
            and options["apply"]
        ):
            raise CommandError(
                "No combines --dry-run "
                "y --apply."
            )


        apply_changes = bool(
            options["apply"]
        )


        excel = Path(
            options["excel"]
        ).expanduser().resolve()


        if not excel.exists():
            raise CommandError(
                f"No existe: {excel}"
            )


        workbook = load_workbook(
            excel,
            data_only=False,
            read_only=False,
        )


        sheet_name = options[
            "sheet"
        ]


        if (
            sheet_name
            not in workbook.sheetnames
        ):
            raise CommandError(
                "No existe la hoja "
                f"{sheet_name}."
            )


        worksheet = workbook[
            sheet_name
        ]


        headers = {
            str(
                worksheet.cell(
                    1,
                    column,
                ).value
                or ""
            ).strip():
                column

            for column in range(
                1,
                worksheet.max_column + 1,
            )
        }


        if "Modelo" not in headers:
            raise CommandError(
                "No existe columna Modelo."
            )


        if "3D" not in headers:
            raise CommandError(
                "No existe columna 3D."
            )


        model_column = (
            headers["Modelo"]
        )

        source_column = (
            headers["3D"]
        )


        technical_records = {
            exact_key(
                item.model_code
            ): item

            for item in (
                AntivibrationTechnicalData
                .objects
                .select_related(
                    "product"
                )
            )
        }


        prepared = []
        errors = []


        for row_number in range(
            2,
            worksheet.max_row + 1,
        ):

            model = str(
                worksheet.cell(
                    row_number,
                    model_column,
                ).value
                or ""
            ).strip()


            if not model:
                continue


            source_url = get_source(
                worksheet.cell(
                    row_number,
                    source_column,
                )
            )


            item = (
                technical_records.get(
                    exact_key(model)
                )
            )


            if not item:
                errors.append(
                    (
                        row_number,
                        model,
                        "No existe en PostgreSQL",
                    )
                )
                continue


            try:

                validate_source_url(
                    source_url
                )

                validate_remote_glb(
                    source_url
                )

            except Exception as exc:

                errors.append(
                    (
                        row_number,
                        model,
                        str(exc),
                    )
                )

                continue


            prepared.append(
                (
                    item,
                    source_url,
                )
            )


            self.stdout.write(
                self.style.SUCCESS(
                    f"{model:<22} OK"
                )
            )


        self.stdout.write("")

        self.stdout.write(
            "=" * 70
        )

        self.stdout.write(
            f"Modelos válidos : "
            f"{len(prepared)}"
        )

        self.stdout.write(
            f"Errores          : "
            f"{len(errors)}"
        )

        self.stdout.write(
            "=" * 70
        )


        if errors:

            for (
                row_number,
                model,
                message,
            ) in errors:

                self.stdout.write(
                    self.style.ERROR(
                        f"Fila {row_number} | "
                        f"{model} | "
                        f"{message}"
                    )
                )


            raise CommandError(
                "No se guardó nada porque "
                "existen errores."
            )


        if len(prepared) != 97:

            raise CommandError(
                "Se esperaban exactamente "
                "97 modelos válidos."
            )


        if not apply_changes:

            self.stdout.write(
                self.style.WARNING(
                    "DRY-RUN: PostgreSQL "
                    "no fue modificado."
                )
            )

            return


        with transaction.atomic():

            for item, source_url in prepared:

                raw_data = dict(
                    item.raw_data
                    or {}
                )

                raw_data[
                    "3D"
                ] = source_url

                item.raw_data = (
                    raw_data
                )

                item.save(
                    update_fields=[
                        "raw_data",
                        "updated_at",
                    ]
                )


        self.stdout.write("")

        self.stdout.write(
            self.style.SUCCESS(
                "97 URL 3D guardadas "
                "correctamente en PostgreSQL."
            )
        )
