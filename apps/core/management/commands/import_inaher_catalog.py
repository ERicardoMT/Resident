from __future__ import annotations

import re
import unicodedata
from html import escape
from pathlib import Path
from typing import Any

from django.core.management.base import (
    BaseCommand,
    CommandError,
)
from django.db import transaction
from openpyxl import load_workbook

from apps.core.models import (
    AntivibrationTechnicalData,
    CatalogCategory,
    CatalogItem,
    CatalogSubcategory,
    LevelerTechnicalData,
)


def normalize(value: Any) -> str:
    """
    Normaliza texto para poder comparar
    encabezados y categorías.
    """

    text = str(
        value or ""
    ).strip()

    text = unicodedata.normalize(
        "NFKD",
        text,
    )

    text = "".join(
        character
        for character in text
        if not unicodedata.combining(
            character
        )
    )

    return (
        text
        .lower()
        .replace("×", "x")
        .strip()
    )


def json_safe_value(value: Any):
    """
    Convierte valores de Excel a tipos
    que PostgreSQL JSONField puede guardar.
    """

    if value is None:
        return ""

    if isinstance(
        value,
        (
            str,
            int,
            float,
            bool,
        ),
    ):
        return value

    return str(value)


def get_value(
    record: dict[str, Any],
    *names: str,
):
    """
    Obtiene un campo independientemente
    de mayúsculas o acentos, respetando
    el orden de prioridad de names.
    """

    normalized_record = {
        normalize(key): value
        for key, value in record.items()
    }

    for name in names:
        normalized_name = normalize(name)

        if normalized_name in normalized_record:
            return normalized_record[
                normalized_name
            ]

    return ""


def parse_capacity_kg(
    value: Any,
) -> float | None:
    """
    Obtiene la capacidad numérica.

    150 KG -> 150
    140 Kg -> 140
    1500 KG -> 1500

    En caso de rango se utiliza el valor
    inferior para ser conservadores.
    """

    text = str(
        value or ""
    ).strip()

    numbers = re.findall(
        r"\d+(?:[.,]\d+)?",
        text,
    )

    if not numbers:
        return None

    values = [
        float(
            number.replace(
                ",",
                ".",
            )
        )
        for number in numbers
    ]

    return min(values)


def find_sheet(
    workbook,
    preferred_names,
):
    """
    Busca primero las hojas conocidas.

    Si no existen, busca una hoja
    que tenga la columna Modelo.
    """

    for name in preferred_names:
        if name in workbook.sheetnames:
            return workbook[name]

    for worksheet in workbook.worksheets:
        headers = [
            normalize(
                worksheet.cell(
                    1,
                    column,
                ).value
            )
            for column in range(
                1,
                worksheet.max_column + 1,
            )
        ]

        if "modelo" in headers:
            return worksheet

    return None


def read_records(
    worksheet,
):
    """
    Lee todas las columnas del Excel.

    No solamente las columnas naranjas.
    """

    headers = []

    for column in range(
        1,
        worksheet.max_column + 1,
    ):
        value = worksheet.cell(
            1,
            column,
        ).value

        if value is None:
            headers.append(None)
        else:
            headers.append(
                str(value).strip()
            )

    records = []

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        record = {}
        has_data = False

        for column, header in enumerate(
            headers,
            start=1,
        ):
            if not header:
                continue

            cell = worksheet.cell(
                row_number,
                column,
            )

            value = cell.value

            # Si la propia celda ya contiene una URL válida,
            # conservamos ese valor.
            #
            # Solo usamos el hipervínculo interno de Excel
            # cuando la celda no contiene directamente una URL.
            if (
                cell.hyperlink
                and normalize(header).startswith(
                    "url"
                )
            ):
                current_value = str(
                    value or ""
                ).strip()

                if not current_value.startswith(
                    (
                        "http://",
                        "https://",
                    )
                ):
                    value = (
                        cell.hyperlink.target
                        or value
                )

            value = json_safe_value(
                value
            )

            if str(value).strip():
                has_data = True

            record[header] = value

        if not has_data:
            continue

        record["_excel_row"] = (
            row_number
        )

        records.append(
            record
        )

    return records


def build_description(
    record: dict[str, Any],
) -> str:
    """
    Genera la ficha técnica HTML.

    Normaliza los nombres de las columnas
    y coloca Imagen y Ficha técnica
    al final, siempre en ese orden.
    """

    rows = []

    image_row = ""
    technical_sheet_row = ""

    field_labels = {
        "imagen": "Imagen",
        "imagen principal (url)": "Imagen",
        "imagen principal": "Imagen",

        "ficha tecnica": "Ficha técnica",
        "url de ficha": "Ficha técnica",
    }

    ignored_fields = {
        "vista previa",
        "vista previa (excel 365 / sheets)",
    }

    for field, value in record.items():

        if field.startswith("_"):
            continue

        normalized_field = normalize(
            field
        )

        # Ignorar columnas auxiliares.
        if normalized_field in ignored_fields:
            continue

        text = str(
            value or ""
        ).strip()

        if not text:
            continue

        display_field = field_labels.get(
            normalized_field,
            str(field),
        )

        safe_field = escape(
            display_field
        )

        safe_value = escape(
            text
        )

        if text.startswith(
            (
                "http://",
                "https://",
            )
        ):
            safe_value = (
                f'<a href="{safe_value}" '
                'target="_blank" '
                'rel="noopener noreferrer">'
                "Ver información"
                "</a>"
            )

        row = (
            "<tr>"
            f"<th>{safe_field}</th>"
            f"<td>{safe_value}</td>"
            "</tr>"
        )

        # Guardamos estas dos filas para
        # colocarlas siempre al final.
        if display_field == "Imagen":
            image_row = row
            continue

        if display_field == "Ficha técnica":
            technical_sheet_row = row
            continue

        rows.append(
            row
        )

    # Orden fijo:
    # 1. Imagen
    # 2. Ficha técnica

    if image_row:
        rows.append(
            image_row
        )

    if technical_sheet_row:
        rows.append(
            technical_sheet_row
        )

    if not rows:
        return ""

    return (
        "<table>"
        "<tbody>"
        + "".join(rows)
        + "</tbody>"
        "</table>"
    )


def leveler_category(
    type_label: str,
):
    """
    Decide si el nivelador pertenece
    a mobiliario o nivelación industrial.
    """

    text = normalize(
        type_label
    )

    if "mobiliario" in text:
        return (
            CatalogCategory.MOBILIARIO
        )

    return (
        CatalogCategory.PATAS_NIVELADORAS
    )


def leveler_subcategory(
    type_label: str,
):
    """
    Usa como subcategoría principal
    la primera característica indicada
    en Tipo.
    """

    first_type = str(
        type_label or ""
    ).split("/")[0]

    text = normalize(
        first_type
    )

    if "alta resistencia" in text:
        return (
            CatalogSubcategory.ALTA_RESISTENCIA
        )

    if "anclaje al piso" in text:
        return (
            CatalogSubcategory.ANCLAJE_PISO
        )

    if "antiderrapante" in text:
        return (
            CatalogSubcategory.ANTIDERRAPANTE
        )

    if (
        "antivibratorio" in text
        or "antivibracion" in text
    ):
        return (
            CatalogSubcategory.ANTIVIBRACION
        )

    if "rotula" in text:
        return (
            CatalogSubcategory.CON_ROTULA
        )

    if "uso rudo" in text:
        return (
            CatalogSubcategory.USO_RUDO
        )

    return ""


def antivibration_subcategory(
    record: dict[str, Any],
):
    """
    Busca una clasificación explícita
    en las columnas no naranjas.

    Si el Excel no la proporciona,
    se deja sin subcategoría.
    """

    classification = get_value(
        record,
        "Subcategoría",
        "Subcategoria",
        "Tipo",
        "Familia",
    )

    text = normalize(
        classification
    )

    if "colgante" in text:
        return (
            CatalogSubcategory.COLGANTES
        )

    if "nivelador" in text:
        return (
            CatalogSubcategory
            .NIVELADORES_MAQUINARIA
        )

    if (
        "anclaje" in text
        and "piso" in text
    ):
        return (
            CatalogSubcategory.SOPORTES_PISO
        )

    if "tacon" in text:
        return (
            CatalogSubcategory.TACONES
        )

    if "pie" in text:
        return (
            CatalogSubcategory.PIES
        )

    return ""


class Command(BaseCommand):
    help = (
        "Importa antivibratorios y niveladores "
        "desde los archivos Excel de INAHER."
    )

    def add_arguments(
        self,
        parser,
    ):
        parser.add_argument(
            "--antivibrators",
            required=True,
            help=(
                "Ruta del Excel de "
                "antivibratorios."
            ),
        )

        parser.add_argument(
            "--levelers",
            required=True,
            help=(
                "Ruta del Excel de "
                "niveladores."
            ),
        )

    def handle(
        self,
        *args,
        **options,
    ):
        antivibration_path = Path(
            options["antivibrators"]
        ).expanduser().resolve()

        leveler_path = Path(
            options["levelers"]
        ).expanduser().resolve()

        if not antivibration_path.exists():
            raise CommandError(
                "No existe el Excel de "
                "antivibratorios: "
                f"{antivibration_path}"
            )

        if not leveler_path.exists():
            raise CommandError(
                "No existe el Excel de "
                "niveladores: "
                f"{leveler_path}"
            )

        with transaction.atomic():

            antivibration_count = (
                self.import_antivibrators(
                    antivibration_path
                )
            )

            leveler_count = (
                self.import_levelers(
                    leveler_path
                )
            )

        self.stdout.write("")

        self.stdout.write(
            self.style.SUCCESS(
                "Importación terminada."
            )
        )

        self.stdout.write(
            f"Antivibratorios procesados: "
            f"{antivibration_count}"
        )

        self.stdout.write(
            f"Niveladores procesados: "
            f"{leveler_count}"
        )

        self.stdout.write(
            f"Productos en catálogo: "
            f"{CatalogItem.objects.count()}"
        )

        self.stdout.write(
            f"Datos antivibratorios: "
            f"{AntivibrationTechnicalData.objects.count()}"
        )

        self.stdout.write(
            f"Datos niveladores: "
            f"{LevelerTechnicalData.objects.count()}"
        )

    def import_antivibrators(
        self,
        excel_path: Path,
    ):
        workbook = load_workbook(
            excel_path,
            data_only=True,
        )

        worksheet = find_sheet(
            workbook,
            [
                "Antivibratorios",
            ],
        )

        if worksheet is None:
            raise CommandError(
                "No se encontró la hoja "
                "de antivibratorios."
            )

        records = read_records(
            worksheet
        )

        imported = 0

        for record in records:

            model_code = str(
                get_value(
                    record,
                    "Modelo",
                )
                or ""
            ).strip()

            if not model_code:
                continue

            capacity_label = str(
                get_value(
                    record,
                    "Capacidad de carga",
                )
                or ""
            ).strip()

            capacity_kg = (
                parse_capacity_kg(
                    capacity_label
                )
            )

            existing_data = (
                AntivibrationTechnicalData
                .objects
                .select_related(
                    "product"
                )
                .filter(
                    model_code=model_code
                )
                .first()
            )

            if existing_data:
                product = (
                    existing_data.product
                )
            else:
                product = (
                    CatalogItem.objects.create(
                        name=model_code,
                        category=(
                            CatalogCategory
                            .ANTIVIBRATORIOS
                        ),
                    )
                )

            product.name = model_code

            product.category = (
                CatalogCategory
                .ANTIVIBRATORIOS
            )

            product.subcategory = (
                antivibration_subcategory(
                    record
                )
                or None
            )

            product.description = (
                build_description(
                    record
                )
            )

            product.price_label = (
                "Cotizar"
            )

            product.is_active = True

            product.sort_order = int(
                record.get(
                    "_excel_row",
                    0,
                )
                or 0
            )

            product.save()

            (
                AntivibrationTechnicalData
                .objects
                .update_or_create(
                    model_code=model_code,
                    defaults={
                        "product":
                            product,

                        "base_diameter":
                            str(
                                get_value(
                                    record,
                                    "Diámetro de base",
                                )
                                or ""
                            ).strip(),

                        "base_height":
                            str(
                                get_value(
                                    record,
                                    "Altura de base",
                                )
                                or ""
                            ).strip(),

                        "screw_diameter":
                            str(
                                get_value(
                                    record,
                                    "Diámetro de tornillo",
                                )
                                or ""
                            ).strip(),

                        "screw_height":
                            str(
                                get_value(
                                    record,
                                    "Altura de tornillo",
                                )
                                or ""
                            ).strip(),

                        "capacity_label":
                            capacity_label,

                        "capacity_kg":
                            capacity_kg,

                        "elastomer_material":
                            str(
                                get_value(
                                    record,
                                    "Material de elastómero",
                                )
                                or ""
                            ).strip(),

                        "screw_material":
                            str(
                                get_value(
                                    record,
                                    "Material de tornillo",
                                )
                                or ""
                            ).strip(),

                        "source_file":
                            excel_path.name,

                        "source_sheet":
                            worksheet.title,

                        "source_row":
                            record.get(
                                "_excel_row"
                            ),

                        "raw_data":
                            record,
                    },
                )
            )

            imported += 1

        return imported

    def import_levelers(
        self,
        excel_path: Path,
    ):
        workbook = load_workbook(
            excel_path,
            data_only=True,
        )

        worksheet = find_sheet(
            workbook,
            [
                "Catálogo niveladores WEB",
            ],
        )

        if worksheet is None:
            raise CommandError(
                "No se encontró la hoja "
                "de niveladores."
            )

        records = read_records(
            worksheet
        )

        imported = 0

        for record in records:

            model_code = str(
                get_value(
                    record,
                    "Modelo",
                )
                or ""
            ).strip()

            if not model_code:
                continue

            type_label = str(
                get_value(
                    record,
                    "Tipo",
                )
                or ""
            ).strip()

            capacity_label = str(
                get_value(
                    record,
                    "Capacidad de carga",
                )
                or ""
            ).strip()

            capacity_kg = (
                parse_capacity_kg(
                    capacity_label
                )
            )

            category = (
                leveler_category(
                    type_label
                )
            )

            subcategory = ""

            if (
                category
                == CatalogCategory
                .PATAS_NIVELADORAS
            ):
                subcategory = (
                    leveler_subcategory(
                        type_label
                    )
                )

            existing_data = (
                LevelerTechnicalData
                .objects
                .select_related(
                    "product"
                )
                .filter(
                    model_code=model_code
                )
                .first()
            )

            if existing_data:
                product = (
                    existing_data.product
                )
            else:
                product = (
                    CatalogItem.objects.create(
                        name=model_code,
                        category=category,
                    )
                )

            product.name = (
                model_code
            )

            product.category = (
                category
            )

            product.subcategory = (
                subcategory
                or None
            )

            product.description = (
                build_description(
                    record
                )
            )

            product.price_label = (
                "Cotizar"
            )

            product.is_active = True

            product.sort_order = int(
                record.get(
                    "_excel_row",
                    0,
                )
                or 0
            )

            product.save()

            product_url = str(
                get_value(
                    record,
                    "URL de ficha",
                )
                or ""
            ).strip()

            (
                LevelerTechnicalData
                .objects
                .update_or_create(
                    model_code=model_code,
                    defaults={
                        "product":
                            product,

                        "capacity_label":
                            capacity_label,

                        "capacity_kg":
                            capacity_kg,

                        "type_label":
                            type_label,

                        "metric_threads":
                            str(
                                get_value(
                                    record,
                                    "Roscas métricas",
                                )
                                or ""
                            ).strip(),

                        "standard_threads":
                            str(
                                get_value(
                                    record,
                                    "Roscas estándar",
                                )
                                or ""
                            ).strip(),

                        "base_diameter":
                            str(
                                get_value(
                                    record,
                                    "Diámetro de base",
                                )
                                or ""
                            ).strip(),

                        "screw_height":
                            str(
                                get_value(
                                    record,
                                    "Altura de tornillo",
                                )
                                or ""
                            ).strip(),

                        "screw_material":
                            str(
                                get_value(
                                    record,
                                    "Material de tornillo",
                                )
                                or ""
                            ).strip(),

                        "base_material":
                            str(
                                get_value(
                                    record,
                                    "Material de base",
                                )
                                or ""
                            ).strip(),

                        "product_url":
                            product_url,

                        "source_file":
                            excel_path.name,

                        "source_sheet":
                            worksheet.title,

                        "source_row":
                            record.get(
                                "_excel_row"
                            ),

                        "raw_data":
                            record,
                    },
                )
            )

            imported += 1

        return imported