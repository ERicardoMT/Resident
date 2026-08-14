from __future__ import annotations

import colorsys
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]

EXCEL_PATH = (
    PROJECT_ROOT
    / "apps"
    / "leveler"
    / "data"
    / (
        "INAHER_Catalogo_Niveladores "
        "revisado 13 de agosto 10_27am.xlsx"
    )
)

OUTPUT_PATH = (
    PROJECT_ROOT
    / "apps"
    / "leveler"
    / "data"
    / "niveladores_campos_naranja.json"
)


def serialize_value(value: Any) -> Any:
    """Convierte valores de Excel en datos compatibles con JSON."""

    if value is None:
        return None

    if isinstance(value, (str, int, float, bool)):
        return value

    if isinstance(value, (date, datetime)):
        return value.isoformat()

    return str(value)


def apply_tint(rgb: tuple[int, int, int], tint: float):
    """Aplica aproximadamente el aclarado u oscurecido de Excel."""

    result = []

    for component in rgb:
        if tint < 0:
            value = component * (1.0 + tint)
        else:
            value = component + (255 - component) * tint

        result.append(
            max(
                0,
                min(255, round(value)),
            )
        )

    return tuple(result)


def load_theme_colors(workbook) -> list[str]:
    """Obtiene los colores del tema utilizado por el libro."""

    if not workbook.loaded_theme:
        return []

    root = ElementTree.fromstring(
        workbook.loaded_theme
    )

    namespace = {
        "a": (
            "http://schemas.openxmlformats.org/"
            "drawingml/2006/main"
        )
    }

    color_scheme = root.find(
        ".//a:clrScheme",
        namespace,
    )

    if color_scheme is None:
        return []

    colors = []

    for item in list(color_scheme):
        child = next(iter(item), None)

        if child is None:
            colors.append("")
            continue

        color = (
            child.attrib.get("val")
            or child.attrib.get("lastClr")
            or ""
        )

        colors.append(color[-6:].upper())

    return colors


def resolve_color(color, theme_colors: list[str]) -> str | None:
    """Convierte colores RGB, indexados o de tema a hexadecimal."""

    if color is None:
        return None

    color_type = color.type
    hex_value = None

    if color_type == "rgb" and color.rgb:
        hex_value = color.rgb[-6:]

    elif color_type == "indexed" and color.indexed is not None:
        index = int(color.indexed)

        if 0 <= index < len(COLOR_INDEX):
            hex_value = COLOR_INDEX[index][-6:]

    elif color_type == "theme" and color.theme is not None:
        index = int(color.theme)

        if 0 <= index < len(theme_colors):
            hex_value = theme_colors[index]

    if not hex_value or len(hex_value) != 6:
        return None

    try:
        rgb = tuple(
            int(
                hex_value[position:position + 2],
                16,
            )
            for position in (0, 2, 4)
        )
    except ValueError:
        return None

    tint = float(color.tint or 0)

    if tint:
        rgb = apply_tint(rgb, tint)

    return "".join(
        f"{component:02X}"
        for component in rgb
    )


def is_orange(hex_color: str | None) -> bool:
    """
    Determina si el color está dentro del rango visual
    correspondiente a naranja.
    """

    if not hex_color:
        return False

    red = int(hex_color[0:2], 16) / 255
    green = int(hex_color[2:4], 16) / 255
    blue = int(hex_color[4:6], 16) / 255

    hue, saturation, value = colorsys.rgb_to_hsv(
        red,
        green,
        blue,
    )

    hue_degrees = hue * 360

    return (
        10 <= hue_degrees <= 55
        and saturation >= 0.25
        and value >= 0.50
    )


def get_cell_fill(cell, theme_colors: list[str]) -> str | None:
    """Obtiene el color de relleno visible de una celda."""

    fill = cell.fill

    if fill.patternType not in ("solid", "darkSolid"):
        return None

    foreground = resolve_color(
        fill.fgColor,
        theme_colors,
    )

    if foreground:
        return foreground

    return resolve_color(
        fill.bgColor,
        theme_colors,
    )


def unique_headers(values: list[Any]) -> list[str]:
    """Evita encabezados vacíos o duplicados."""

    headers = []
    repetitions: dict[str, int] = {}

    for index, value in enumerate(values, start=1):
        header = str(
            value or f"campo_{index}"
        ).strip()

        repetitions[header] = (
            repetitions.get(header, 0) + 1
        )

        if repetitions[header] > 1:
            header = (
                f"{header}_"
                f"{repetitions[header]}"
            )

        headers.append(header)

    return headers


def find_orange_header_row(
    worksheet,
    theme_colors: list[str],
):
    """
    Busca la fila que contiene la mayor cantidad
    de encabezados naranjas.
    """

    candidates = []

    max_rows = min(
        worksheet.max_row,
        80,
    )

    for row_number in range(1, max_rows + 1):
        orange_cells = []

        for cell in worksheet[row_number]:
            if cell.value in (None, ""):
                continue

            color = get_cell_fill(
                cell,
                theme_colors,
            )

            if is_orange(color):
                orange_cells.append(
                    {
                        "column": cell.column,
                        "coordinate": cell.coordinate,
                        "value": cell.value,
                        "color": color,
                    }
                )

        if orange_cells:
            candidates.append(
                {
                    "row": row_number,
                    "cells": orange_cells,
                }
            )

    if not candidates:
        return None

    candidates.sort(
        key=lambda candidate: (
            -len(candidate["cells"]),
            candidate["row"],
        )
    )

    return candidates[0]


def extract_worksheet(
    worksheet,
    header_information,
):
    """Extrae únicamente las columnas con encabezados naranjas."""

    header_row = header_information["row"]
    orange_cells = header_information["cells"]

    columns = [
        cell["column"]
        for cell in orange_cells
    ]

    raw_headers = [
        worksheet.cell(
            row=header_row,
            column=column,
        ).value
        for column in columns
    ]

    headers = unique_headers(raw_headers)

    fields = []

    for header, cell_info in zip(
        headers,
        orange_cells,
        strict=True,
    ):
        fields.append(
            {
                "name": header,
                "column": get_column_letter(
                    cell_info["column"]
                ),
                "coordinate": cell_info["coordinate"],
                "color": cell_info["color"],
            }
        )

    records = []

    for row_number in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):
        values = [
            serialize_value(
                worksheet.cell(
                    row=row_number,
                    column=column,
                ).value
            )
            for column in columns
        ]

        if all(
            value in (None, "")
            for value in values
        ):
            continue

        record = dict(
            zip(
                headers,
                values,
                strict=True,
            )
        )

        record["_excel_row"] = row_number

        records.append(record)

    return {
        "sheet": worksheet.title,
        "header_row": header_row,
        "fields": fields,
        "records": records,
    }


def main():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(
            (
                "No se encontró el archivo:\n"
                f"{EXCEL_PATH}"
            )
        )

    workbook = load_workbook(
        EXCEL_PATH,
        data_only=True,
        read_only=False,
    )

    theme_colors = load_theme_colors(
        workbook
    )

    extracted_sheets = []

    print(f"Archivo: {EXCEL_PATH.name}")
    print(f"Hojas: {', '.join(workbook.sheetnames)}")
    print()

    for worksheet in workbook.worksheets:
        header_information = find_orange_header_row(
            worksheet,
            theme_colors,
        )

        if header_information is None:
            print(
                (
                    f'Hoja "{worksheet.title}": '
                    "sin encabezados naranjas."
                )
            )
            continue

        result = extract_worksheet(
            worksheet,
            header_information,
        )

        extracted_sheets.append(result)

        print(f'Hoja: "{worksheet.title}"')
        print(
            (
                "Fila de encabezados detectada: "
                f'{result["header_row"]}'
            )
        )
        print("Campos naranjas detectados:")

        for field in result["fields"]:
            print(
                (
                    f'  - {field["column"]}: '
                    f'{field["name"]} '
                    f'#{field["color"]}'
                )
            )

        print(
            (
                "Registros extraídos: "
                f'{len(result["records"])}'
            )
        )

        print("Primeros 3 registros:")

        for record in result["records"][:3]:
            print(
                json.dumps(
                    record,
                    ensure_ascii=False,
                )
            )

        print()

    if not extracted_sheets:
        raise RuntimeError(
            (
                "No se detectaron encabezados con "
                "relleno naranja en ninguna hoja."
            )
        )

    payload = {
        "source_file": EXCEL_PATH.name,
        "selection_rule": (
            "Únicamente columnas con encabezado naranja"
        ),
        "sheets": extracted_sheets,
    }

    OUTPUT_PATH.write_text(
        json.dumps(
            payload,
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print("Archivo generado:")
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
    