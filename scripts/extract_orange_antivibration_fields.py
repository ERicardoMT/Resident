from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent.parent


def get_fill_description(cell) -> str:
    """
    Devuelve una descripción legible del color
    de relleno utilizado por Excel.
    """

    color = cell.fill.fgColor

    if color.type == "rgb":
        rgb = str(
            color.rgb or ""
        ).upper()

        if len(rgb) >= 6:
            rgb = rgb[-6:]

        return f"RGB-{rgb}"

    if color.type == "theme":
        return (
            f"THEME-{color.theme}"
            f"-TINT-{color.tint}"
        )

    if color.type == "indexed":
        return (
            f"INDEXED-{color.indexed}"
        )

    return str(
        color.type or ""
    )


def is_orange(cell) -> bool:
    """
    Detecta el naranja utilizado por los
    catálogos de INAHER.

    Puede venir guardado como:
    - RGB F79646
    - Accent 6 del tema de Excel (theme 9)
    """

    if cell.fill.fill_type is None:
        return False

    color = cell.fill.fgColor

    if color.type == "rgb":
        rgb = str(
            color.rgb or ""
        ).upper()

        return rgb.endswith(
            "F79646"
        )

    if color.type == "theme":
        return color.theme == 9

    return False


def find_header_row(ws) -> int | None:
    """
    Busca en las primeras filas aquella
    que tenga más celdas naranjas.
    """

    best_row = None
    best_count = 0

    maximum_row = min(
        ws.max_row,
        30,
    )

    for row_number in range(
        1,
        maximum_row + 1,
    ):
        orange_count = 0

        for cell in ws[row_number]:
            if (
                cell.value is not None
                and is_orange(cell)
            ):
                orange_count += 1

        if orange_count > best_count:
            best_count = orange_count
            best_row = row_number

    if best_count == 0:
        return None

    return best_row


def print_colored_cells(ws):
    """
    Diagnóstico: muestra las celdas con
    relleno de color de las primeras filas.
    """

    print()
    print(
        f'Colores encontrados en "{ws.title}":'
    )

    found = False

    maximum_row = min(
        ws.max_row,
        15,
    )

    for row_number in range(
        1,
        maximum_row + 1,
    ):
        for cell in ws[row_number]:
            if cell.value is None:
                continue

            if cell.fill.fill_type is None:
                continue

            found = True

            print(
                f"  {cell.coordinate}: "
                f"{cell.value} "
                f"[{get_fill_description(cell)}]"
            )

    if not found:
        print(
            "  No se encontraron celdas "
            "con relleno en las primeras "
            "15 filas."
        )


def extract_sheet(ws):
    header_row = find_header_row(
        ws
    )

    if header_row is None:
        print_colored_cells(
            ws
        )

        print()
        print(
            f'Hoja "{ws.title}": '
            "sin encabezados naranjas "
            "detectados."
        )

        return None

    orange_columns = []

    for cell in ws[header_row]:
        if (
            cell.value is not None
            and is_orange(cell)
        ):
            orange_columns.append(
                (
                    cell.column,
                    str(
                        cell.value
                    ).strip(),
                    get_fill_description(
                        cell
                    ),
                )
            )

    print()
    print(
        f'Hoja: "{ws.title}"'
    )

    print(
        "Fila de encabezados detectada:",
        header_row,
    )

    print(
        "Campos naranjas detectados:"
    )

    for (
        column,
        name,
        color,
    ) in orange_columns:

        letter = ws.cell(
            header_row,
            column,
        ).column_letter

        print(
            f"  - {letter}: "
            f"{name} [{color}]"
        )

    records = []

    for row_number in range(
        header_row + 1,
        ws.max_row + 1,
    ):
        record = {}
        has_data = False

        for (
            column,
            name,
            _,
        ) in orange_columns:

            value = ws.cell(
                row_number,
                column,
            ).value

            if value is not None:
                has_data = True

            record[name] = (
                value
                if value is not None
                else ""
            )

        if not has_data:
            continue

        record[
            "_excel_row"
        ] = row_number

        records.append(
            record
        )

    print(
        "Registros extraídos:",
        len(records),
    )

    print(
        "Primeros 3 registros:"
    )

    for record in records[:3]:
        print(
            json.dumps(
                record,
                ensure_ascii=False,
                default=str,
            )
        )

    return {
        "sheet": ws.title,
        "header_row": header_row,
        "orange_fields": [
            name
            for (
                _,
                name,
                _,
            )
            in orange_columns
        ],
        "records": records,
    }


def main():
    if len(sys.argv) < 2:
        print(
            "Uso:"
        )

        print(
            "python "
            "scripts/"
            "extract_orange_antivibration_fields.py "
            "\"archivo.xlsx\""
        )

        raise SystemExit(1)

    excel_path = Path(
        sys.argv[1]
    ).expanduser().resolve()

    if not excel_path.exists():
        print(
            "No existe el archivo:",
            excel_path,
        )

        raise SystemExit(1)

    workbook = load_workbook(
        excel_path,
        data_only=True,
    )

    print(
        "Archivo:",
        excel_path.name,
    )

    print(
        "Hojas:",
        ", ".join(
            workbook.sheetnames
        ),
    )

    extracted_sheets = []

    for ws in workbook.worksheets:
        result = extract_sheet(
            ws
        )

        if result:
            extracted_sheets.append(
                result
            )

    output_directory = (
        PROJECT_ROOT
        / "apps"
        / "stops"
        / "data"
    )

    output_directory.mkdir(
        parents=True,
        exist_ok=True,
    )

    output_path = (
        output_directory
        / (
            "antivibratorios_"
            "campos_naranja.json"
        )
    )

    output_data = {
        "source_file": (
            excel_path.name
        ),
        "sheets": (
            extracted_sheets
        ),
    }

    output_path.write_text(
        json.dumps(
            output_data,
            ensure_ascii=False,
            indent=2,
            default=str,
        ),
        encoding="utf-8",
    )

    print()
    print(
        "Archivo generado:"
    )

    print(
        output_path
    )


if __name__ == "__main__":
    main()